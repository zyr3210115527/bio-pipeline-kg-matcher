#!/usr/bin/env python3
"""用配置的 LLM API 实跑 benchmark（96 例 + 150 例）。
- 96 例：评 pipeline Top1/Top3 + 数据文件精确匹配（多一少一都算错）。
- 150 例：仅评 pipeline（含多工具题，按集合是否覆盖判定）。
"""
import os, sys, json, warnings, re
warnings.filterwarnings("ignore")

os.environ["FORCE_RULE"] = "0"
os.environ["LLM_MODE"] = "api"
if not (os.environ.get("LLM_API_KEY") or os.environ.get("DEEPSEEK_API_KEY")):
    raise RuntimeError("Set LLM_API_KEY or DEEPSEEK_API_KEY before running the benchmark")
os.environ.setdefault("LLM_BASE_URL", "https://api.deepseek.com/chat/completions")
os.environ.setdefault("LLM_MODEL", "deepseek-v4-pro")
os.environ.setdefault("LLM_TIMEOUT", "60")
os.environ["LLM_TIMEOUT"] = "60"

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import openpyxl
from pipeline_router import route_pipeline_request

BENCH96 = "/Users/zhouyiran/data/progress/0709/incoming/96例问题-数据-工具对应表.xlsx"
BENCH150 = "/Users/zhouyiran/data/progress/0709/incoming/问题与调用工具150.xlsx"

def split_set(cell):
    if not cell:
        return set()
    return {p.strip() for p in re.split(r"[\n;,、]+", str(cell)) if p.strip()}

def norm_tool(t):
    # 归一化 her3/4/5/6_pfs_survival 这类标注笔误 → her2_pfs_survival
    return re.sub(r"her[3-9]_pfs_survival", "her2_pfs_survival", t.strip())

def pred_file_set(result):
    files = (result.get("matched_data") or {}).get("file_candidates") or []
    return {str(f.get("files")).strip() for f in files if f.get("files")}


def run96():
    wb = openpyxl.load_workbook(BENCH96, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = [r for r in list(ws.iter_rows(values_only=True))[1:] if r[1]]
    wb.close()
    n = len(rows)
    top1=top3=dexact=e2e=used=0
    recs=[]
    for i,r in enumerate(rows,1):
        q, gdata, gtool = r[1], r[2], r[3]
        gtools = {norm_tool(t) for t in split_set(gtool)}
        gprimary = list(gtools)[0] if gtools else ""
        gfiles = split_set(gdata)
        res = route_pipeline_request(q, top_k=5)["result"]
        mp = res.get("matched_pipelines") or []
        pids = [p["pipeline_id"] for p in mp]
        p1 = pids[0] if pids else None
        if (res.get("llm_metadata") or {}).get("used"): used+=1
        t1 = p1 in gtools
        t3 = bool(gtools & set(pids[:3]))
        pf = pred_file_set(res)
        de = (pf==gfiles) and bool(gfiles)
        top1+=t1; top3+=t3; dexact+=de
        if t1 and de: e2e+=1
        recs.append({"idx":i,"q":q,"gold":gprimary,"pred_top1":p1,"pred_top3":pids[:3],
                     "top1":t1,"top3":t3,"gold_files":sorted(gfiles),"pred_files":sorted(pf),"data_exact":de})
        print(f"96 [{i:>2}/{n}] tool{'✓' if t1 else '✗'} data{'✓' if de else '✗'} gold={gprimary:<30} pred={p1}",flush=True)
    return {"n":n,"used":used,"top1":top1,"top3":top3,"data_exact":dexact,"e2e":e2e,"records":recs}


def run150():
    wb = openpyxl.load_workbook(BENCH150, read_only=True, data_only=True)
    ws = wb["Sheet4"]
    rows = [r for r in list(ws.iter_rows(values_only=True))[1:] if r[1]]
    wb.close()
    n=len(rows)
    top1=top3=exact_set=used=0
    recs=[]
    for i,r in enumerate(rows,1):
        q, gtool = r[1], r[2]
        gtools = {norm_tool(t) for t in split_set(gtool)}
        res = route_pipeline_request(q, top_k=5)["result"]
        mp = res.get("matched_pipelines") or []
        pids = [p["pipeline_id"] for p in mp]
        p1 = pids[0] if pids else None
        if (res.get("llm_metadata") or {}).get("used"): used+=1
        # Top1: 预测第一个属于标准工具集合
        t1 = p1 in gtools
        # Top3: 标准集合与前三有交集
        t3 = bool(gtools & set(pids[:3]))
        # 集合精确(多工具题)：标准工具集合 ⊆ 预测集合 且 预测未超出（按标准集合大小截取预测）
        es = gtools == set(pids[:len(gtools)]) if len(gtools)>1 else (p1 in gtools)
        top1+=t1; top3+=t3; exact_set+=es
        recs.append({"idx":i,"q":q,"gold":sorted(gtools),"pred_top1":p1,"pred_top3":pids[:3],
                     "top1":t1,"top3":t3,"set_exact":es})
        print(f"150[{i:>3}/{n}] tool{'✓' if t1 else '✗'} gold={','.join(sorted(gtools)):<40} pred={p1}",flush=True)
    return {"n":n,"used":used,"top1":top1,"top3":top3,"set_exact":exact_set,"records":recs}


if __name__ == "__main__":
    which = sys.argv[1] if len(sys.argv)>1 else "both"
    out={}
    if which in ("96","both"):
        out["bench96"]=run96()
    if which in ("150","both"):
        out["bench150"]=run150()
    with open("bench_deepseek_full.json","w",encoding="utf-8") as f:
        json.dump(out,f,ensure_ascii=False,indent=2)
    print("\n"+"="*64)
    if "bench96" in out:
        b=out["bench96"]; n=b["n"]
        print(f"[96 例]  LLM生效 {b['used']}/{n}")
        print(f"  Pipeline Top1        : {b['top1']}/{n} = {b['top1']/n*100:.2f}%")
        print(f"  Pipeline Top3        : {b['top3']}/{n} = {b['top3']/n*100:.2f}%")
        print(f"  数据精确匹配(多一少一算错): {b['data_exact']}/{n} = {b['data_exact']/n*100:.2f}%")
        print(f"  End-to-end(Top1且数据精确): {b['e2e']}/{n} = {b['e2e']/n*100:.2f}%")
    if "bench150" in out:
        b=out["bench150"]; n=b["n"]
        print(f"[150 例] LLM生效 {b['used']}/{n}")
        print(f"  Pipeline Top1        : {b['top1']}/{n} = {b['top1']/n*100:.2f}%")
        print(f"  Pipeline Top3        : {b['top3']}/{n} = {b['top3']/n*100:.2f}%")
        print(f"  多工具集合精确匹配    : {b['set_exact']}/{n} = {b['set_exact']/n*100:.2f}%")
    print("明细 → bench_deepseek_full.json")
